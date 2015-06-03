from reporting.models import *
import openpyxl
import os
from diplome.settings import BASE_DIR, STATIC_URL


FILE_NAME_TEMPLATE = os.path.join(os.path.dirname(__file__), 'templates/')
FILE_NAME_DESTINATION = os.path.join(BASE_DIR, 'static', 'reductions/')
PASS_TYPES = [passType[0] for passType in PASS_TYPES]
MARITAL_STATUSES = {
    'single': 'Не одружений',
    'married': 'Одружений'
}
MONTHS = {
    '01': 'Січень',
    '02': 'Лютий',
    '03': 'Березень',
    '04': 'Квітень',
    '05': 'Травень',
    '06': 'Червень',
    '07': 'Липень',
    '08': 'Серпень',
    '09': 'Вересень',
    '10': 'Жовтень',
    '11': 'Листопад',
    '12': 'Грудень',
}


def generate_journal_excel(group):
    sheets = {
        'а) Обгортка': fill_start_page,
        'б) Розподіл громад. доручень': fill_public_orders,
        'в) Відомості': fill_news,
        'в2) Інформація для мене': fill_info_for_me,
        'г) План громадсько-вих. роботи': fill_plan,
        'д) Облік відвідувань': fill_accounting,
        'е) Громад.-Вих. заходи': fill_public_attitude,
        'ж) Індивід. робота з студентами': fill_individual_work,
        'з) Звіт про гром.-вих. роботу': fill_report
    }
    workbook = openpyxl.load_workbook(FILE_NAME_TEMPLATE + 'journal.xlsx')
    group = Group.objects.get(id=group)
    students = Student.objects.filter(group=group)

    for sheet in sheets:
        ws = workbook.get_sheet_by_name(sheet)
        sheets[sheet](ws, group)

    workbook.save(FILE_NAME_DESTINATION + 'journal.xlsx')
    return STATIC_URL + 'reductions/' + 'journal.xlsx'


def fill_start_page(ws, group):
    ws['B6'] = ' '.join(['Факультет:', group.department.title])
    ws['B7'] = ' '.join(['Куратор:', group.curator.lastName, group.curator.firstName, group.curator.middleName])
    ws['B8'] = ' '.join(['Староста:', group.leader.lastName, group.leader.firstName])


def fill_public_orders(ws, group):
    ws['A2'] = '1.Курс'
    ws['B2'] = group.yearStudy
    ws['A3'] = '2.Група'
    ws['B3'] = group.number
    ws['A4'] = '3.Староста'
    ws['B4'] = group.leader.lastName + ' ' + group.leader.firstName if group.leader else ''
    ws['A5'] = '4.Заступник старости'
    ws['B5'] = group.deputyHeadman.lastName + ' ' + group.deputyHeadman.firstName if group.deputyHeadman else ''
    ws['A6'] = '5.Профгрупорг'
    ws['B6'] = group.organizer.lastName + ' ' + group.organizer.firstName if group.organizer else ''
    ws['A7'] = '6.Культурно-дозвіллєва робота'
    ws['B7'] = group.culturalWork.lastName + ' ' + group.culturalWork.firstName if group.culturalWork else ''
    ws['A8'] = '7.Спортивно-оздоровча робота'
    ws['B8'] = group.healthWork.lastName + ' ' + group.healthWork.firstName if group.healthWork else ''
    ws['A9'] = '8.Редколегія'
    start_index = 9
    for person in group.editorialBoard.all():
        ws['B' + str(start_index)] = person.lastName + ' ' + person.firstName
        start_index += 1
    ws['A' + str(start_index)] = '9.Інші доручення'
    for person in group.otherTasks.all():
        ws['B' + str(start_index)] = person.lastName + ' ' + person.firstName
        start_index += 1


def fill_news(ws, group):
    students = group.student_set.all()
    start_index = 4
    for index, student in enumerate(students):
        ws['A' + str(start_index)] = index + 1
        ws['B' + str(start_index)] = ' '.join([student.lastName, student.firstName, student.middleName])
        ws['C' + str(start_index)] = student.dateBirth
        ws['D' + str(start_index)] = student.nationality or '~'
        ws['E' + str(start_index)] = MARITAL_STATUSES[student.maritalStatus]
        ws['F' + str(start_index)] = student.id
        ws['G' + str(start_index)] = ' '.join([student.address, 'тел.', student.phone])
        ws['H' + str(start_index)] = ' '.join([student.father.fullName, student.father.position + ',', student.mother.fullName, student.mother.position ])
        start_index += 1


def fill_info_for_me(ws, group):
    students = group.student_set.all()
    ws['A1'] = group.number
    start_index = 3
    for index, student in enumerate(students):
        ws['A' + str(start_index)] = index + 1
        ws['B' + str(start_index)] = ' '.join([student.lastName, student.firstName, student.middleName])
        ws['C' + str(start_index)] = student.additional
        ws['D' + str(start_index)] = student.additional
        ws['E' + str(start_index)] = student.phone
        ws['F' + str(start_index)] = student.email
        ws['G' + str(start_index)] = student.address
        ws['H' + str(start_index)] = '\n'.join([student.father.phone, student.mother.phone])
        ws['I' + str(start_index)] = student.address
        ws['J' + str(start_index)] = student.school
        start_index += 1


def fill_plan(ws, group):
    start_index = 9
    public_plans = PublicPlan.objects.filter(group=group)
    ws['C2'] = 'Завідувач кафедри ' + group.department.headOfDepartment.lastName + ' ' + group.department.headOfDepartment.firstName + ' ' + group.department.headOfDepartment.middleName if group.department.headOfDepartment else ''

    for index, plan in enumerate(public_plans):
        ws['A' + str(start_index)] = index + 1
        ws['B' + str(start_index)] = plan.event
        ws['C' + str(start_index)] = plan.date
        ws['D' + str(start_index)] = plan.responsive
        ws['E' + str(start_index)] = plan.description
        start_index += 1
    ws['C54'] = group.curator.lastName + ' ' + group.curator.firstName + ' ' + group.curator.middleName if group.curator else ''


def fill_accounting(ws, group):
    start_index = 5
    students = Student.objects.filter(group=group)

    for index, student in enumerate(students):
        passes_first_semester = student.pass_set.filter(class_passed__semester=1)
        ws['A' + str(start_index)] = index + 1
        ws['B' + str(start_index)] = ' '.join([student.lastName, student.firstName, student.middleName])

        cells = {
            'pass': {
                9: 'D',
                10: 'F',
                11: 'H',
                12: 'J',
                1: 'L'
            },
            'other': {
                9: 'C',
                10: 'E',
                11: 'G',
                12: 'I',
                1: 'K'
            },
        }

        for month in [9, 10, 11, 12, 1]:
            passes_no_reason = passes_first_semester.filter(date__month=month, type='pass')
            passes_others = passes_first_semester.filter(date__month=month).exclude(type='pass')
            ws[cells['pass'][month] + str(start_index)] = passes_no_reason.count() * 2
            ws[cells['other'][month] + str(start_index)] = passes_others.count() * 2
        start_index += 1

    start_index = 5
    for index, student in enumerate(students):
        passes_second_semester = student.pass_set.filter(class_passed__semester=2)
        ws['M' + str(start_index)] = index + 1
        ws['N' + str(start_index)] = ' '.join([student.lastName, student.firstName, student.middleName])

        cells = {
            'pass': {
                2: 'P',
                3: 'R',
                4: 'T',
                5: 'V',
                6: 'X'
            },
            'other': {
                2: 'O',
                3: 'Q',
                4: 'S',
                5: 'U',
                6: 'W'
            }
        }

        for month in [2, 3, 4, 5, 6]:
            passes_no_reason = passes_second_semester.filter(date__month=month, type='pass')
            passes_others = passes_second_semester.filter(date__month=month).exclude(type='pass')
            ws[cells['pass'][month] + str(start_index)] = passes_no_reason.count() * 2
            ws[cells['other'][month] + str(start_index)] = passes_others.count() * 2
        start_index += 1


def fill_public_attitude(ws, group):
    start_index = 3
    public_plans_first_semester = PublicPlan.objects.filter(semester=1)
    public_plans_second_semester = PublicPlan.objects.filter(semester=2)

    for index, plan in enumerate(public_plans_first_semester):
        ws['A' + str(start_index)] = index + 1
        ws['B' + str(start_index)] = plan.date
        ws['C' + str(start_index)] = plan.event
        ws['D' + str(start_index)] = plan.amountHours
        ws['E' + str(start_index)] = plan.amountPresent
        start_index += 1

    start_index = 3
    for index, plan in enumerate(public_plans_second_semester):
        ws['F' + str(start_index)] = index + 1
        ws['G' + str(start_index)] = plan.date
        ws['H' + str(start_index)] = plan.event
        ws['I' + str(start_index)] = plan.amountHours
        ws['J' + str(start_index)] = plan.amountPresent
        start_index += 1


def fill_individual_work(ws, group):
    works = StudentWork.objects.filter(group=group)
    start_index = 2

    for index, work in enumerate(works):
        ws['A' + str(start_index)] = work.text


def fill_report(ws, group):
    plans = PublicPlan.objects.filter(group=group)
    start_index = 5

    for index, plan in enumerate(plans):
        ws['A' + str(start_index)] = plan.event + ' (' + plan.date + ')'



