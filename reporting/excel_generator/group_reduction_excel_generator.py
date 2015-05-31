from reporting.models import Student, Pass, Group, Faculty, PASS_TYPES
import openpyxl
import os
from diplome.settings import BASE_DIR, STATIC_URL
from django.db.models import Count
from datetime import date


FILE_NAME_TEMPLATE = os.path.join(os.path.dirname(__file__), 'templates/')
FILE_NAME_DESTINATION = os.path.join(BASE_DIR, 'static', 'reductions/')
PASS_TYPES = [passType[0] for passType in PASS_TYPES]
MONTHS = {
    '01': 'Січень',
    '02': 'Лютий',
    '03': 'Березень',
    '04': 'Квітень',
    '05': 'Травень',
    '06': 'Червень',
    '07': 'Липень',
    '08': 'Серпень',
    '09': 'Вересень',
    '10': 'Жовтень',
    '11': 'Листопад',
    '12': 'Грудень',
}
MONTH_PASSES_CELLS = {
    'pass': 'F',
    'sickness': 'C',
    'statement': 'D',
    'watch': 'E',
    'total': 'G'
}
SEMESTER_PASSES_CELLS = {
    'pass': 'O',
    'sickness': 'L',
    'statement': 'M',
    'watch': 'N',
    'total': 'P'
}

PASSES_PER_MONTH_BY_CLASS_TYPE_CELLS = {
    'lecture_total': 'H',
    'lecture': 'I',
    'practice_total': 'J',
    'practice': 'K'
}

PASSES_PER_SEMESTER_BY_CLASS_TYPE_CELLS = {
    'lecture_total': 'Q',
    'lecture': 'R',
    'practice_total': 'S',
    'practice': 'T'
}


def generate_group_reduction(faculty, group, month, year):
    workbook = openpyxl.load_workbook(FILE_NAME_TEMPLATE + 'zvedenia1.xlsx')
    faculty = Faculty.objects.get(id=faculty)
    group = Group.objects.get(id=group)
    students = Student.objects.filter(group=group)
    ws = workbook.active

    ws['A2'] = 'про пропуски занять студентами факультету ' + faculty.title
    ws['A3'] = 'за ' + MONTHS[month]
    ws['C5'] = 'за ' + MONTHS[month]
    ws['C3'] = ' місяць ' + str(year) + ' навчального року'
    ws['A4'] = 'Група ' + str(group.number)

    start_index = 9
    for index, student in enumerate(students):
        start_index += 1
        ws['A' + str(start_index)] = index + 1
        ws['B' + str(start_index)] = student.full_name()
        start_semester_date = faculty.startFirstSemester if int(month) >= 9 else faculty.startSecondSemester

        passes_per_month = student.pass_set.filter(date__year=year, date__month=month)
        total_passes_per_month = passes_per_month.count()
        grouped_passes_per_month = passes_per_month.values('type').annotate(count=Count('type'))

        passes_per_semester = student.pass_set.filter(date__year=year, date__gte=start_semester_date, date__lt=date(int(year), int(month) + 1, 1))
        total_passes_per_semester = passes_per_semester.count()
        grouped_passes_per_semester = passes_per_semester.values('type').annotate(count=Count('type'))

        passes_per_month_grouped_by_class_type = passes_per_month.filter(type='pass').values('class_passed__type').annotate(count=Count('class_passed__type'))
        passes_per_semester_grouped_by_class_type = passes_per_semester.filter(type='pass').values('class_passed__type').annotate(count=Count('class_passed__type'))

        for passItem in grouped_passes_per_month:
            ws[MONTH_PASSES_CELLS[passItem['type']] + str(start_index)] = passItem['count'] * 2
            ws[MONTH_PASSES_CELLS['total'] + str(start_index)] = total_passes_per_month * 2

        for passItem in grouped_passes_per_semester:
            ws[SEMESTER_PASSES_CELLS[passItem['type']] + str(start_index)] = passItem['count'] * 2
            ws[SEMESTER_PASSES_CELLS['total'] + str(start_index)] = total_passes_per_semester * 2

        for passItem in passes_per_month_grouped_by_class_type:
            ws[PASSES_PER_MONTH_BY_CLASS_TYPE_CELLS[passItem['class_passed__type']] + str(start_index)] = passItem['count'] * 2
            ws[PASSES_PER_MONTH_BY_CLASS_TYPE_CELLS[passItem['class_passed__type'] + '_' + 'total'] + str(start_index)] = passes_per_month.filter(class_passed__type=passItem['class_passed__type']).count() * 2

        for passItem in passes_per_semester_grouped_by_class_type:
            ws[PASSES_PER_SEMESTER_BY_CLASS_TYPE_CELLS[passItem['class_passed__type']] + str(start_index)] = passItem['count'] * 2
            ws[PASSES_PER_SEMESTER_BY_CLASS_TYPE_CELLS[passItem['class_passed__type'] + '_' + 'total'] + str(start_index)] = passes_per_semester.filter(class_passed__type=passItem['class_passed__type']).count() * 2

    passes_per_month = Pass.objects.filter(date__year=year, date__month=month)
    passes_per_semester = Pass.objects.filter(date__year=year, date__gte=start_semester_date, date__lt=date(int(year), int(month) + 1, 1))
    ws['L51'] = passes_per_month.distinct('student').count()
    ws['L52'] = passes_per_month.filter(type='pass').distinct('student').count()
    ws['L53'] = passes_per_month.filter(type='pass').count() * 2
    ws['L54'] = passes_per_month.all().count() * 2

    ws['R51'] = passes_per_semester.distinct('student').count()
    ws['R52'] = passes_per_semester.filter(type='pass').distinct('student').count()
    ws['R53'] = passes_per_semester.filter(type='pass').count() * 2
    ws['R54'] = passes_per_semester.all().count() * 2

    workbook.save(FILE_NAME_DESTINATION + 'group_reduction_per_month.xlsx')
    return STATIC_URL + 'reductions/' + 'group_reduction_per_month.xlsx'
