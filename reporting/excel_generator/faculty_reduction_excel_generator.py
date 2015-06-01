from reporting.models import Department, Pass, Group, Faculty, Student, Class, PASS_TYPES
import openpyxl
import os
from diplome.settings import BASE_DIR, STATIC_URL
from django.db.models import Count
from datetime import date


FILE_NAME_TEMPLATE = os.path.join(os.path.dirname(__file__), 'templates/')
FILE_NAME_DESTINATION = os.path.join(BASE_DIR, 'static', 'reductions/')
PASS_TYPES = [passType[0] for passType in PASS_TYPES]
START_LINE = 12
MONTHS = {
    '01': 'Січень',
    '02': 'Лютий',
    '03': 'Березень',
    '04': 'Квітень',
    '05': 'Травень',
    '06': 'Червень',
    '07': 'Липень',
    '08': 'Серпень',
    '09': 'Вересень',
    '10': 'Жовтень',
    '11': 'Листопад',
    '12': 'Грудень',
}


def generate_faculty_reduction_per_month(faculty, month, year, additional):
    workbook = openpyxl.load_workbook(FILE_NAME_TEMPLATE + 'zvedenia2.xlsx')
    faculty = Faculty.objects.get(id=faculty)
    start_semester_date = faculty.startFirstSemester if int(month) >= 9 else faculty.startSecondSemester
    departments = Department.objects.all()
    group = Group.objects.all()
    ws = workbook.active

    ws['A4'] = 'факультету ' + faculty.title
    ws['A6'] = ' '.join(['за', MONTHS[month], str(year), 'навчального року'])

    start_index = START_LINE
    for index, department in enumerate(departments):
        start_index += 1
        ws['A' + str(start_index)] = index + 1
        ws['B' + str(start_index)] = department.title
        groups = department.group_set.all()
        students = Student.objects.filter(group__id__in=groups)
        classes = Class.objects.filter(group__id__in=groups)
        all_passes = Pass.objects.filter(student__id__in=students, date__year=year, date__gte=start_semester_date, date__lt=date(int(year), int(month) + 1, 1))
        fill_line(all_passes, students, classes, start_index, ws, additional, None, department)
        start_index = fill_department(year, start_semester_date, date(int(year), int(month) + 1, 1), groups, start_index, ws, additional, department)

    groups = Group.objects.filter(department__faculty=faculty)
    students = Student.objects.filter(group__id__in=groups)
    classes = Class.objects.filter(group__id__in=groups)
    all_passes = Pass.objects.filter(student__id__in=students, date__year=year, date__gte=start_semester_date, date__lt=date(int(year), int(month) + 1, 1))
    start_index += 2
    ws['B' + str(start_index)] = 'Всього по факультету:'
    fill_line(all_passes, students, classes, start_index, ws, additional)
    fill_department(year, start_semester_date, date(int(year), int(month) + 1, 1), groups, start_index, ws, additional)

    workbook.save(FILE_NAME_DESTINATION + 'faculty_reduction_per_month.xlsx')
    return STATIC_URL + 'reductions/' + 'faculty_reduction_per_month.xlsx'


def generate_faculty_reduction_per_semester(faculty, semester, year, additional):
    workbook = openpyxl.load_workbook(FILE_NAME_TEMPLATE + 'zvedenia2.xlsx')
    faculty = Faculty.objects.get(id=faculty)
    start_semester_date = faculty.startFirstSemester if int(semester) == 1 else faculty.startSecondSemester
    end_semester_date = faculty.endFirstSemester if int(semester) == 1 else faculty.endSecondSemester
    departments = Department.objects.all()
    group = Group.objects.all()
    ws = workbook.active

    ws['A4'] = 'факультету ' + faculty.title
    ws['A6'] = ' '.join(['за', semester, 'семестер', str(year), 'навчального року'])

    start_index = START_LINE
    for index, department in enumerate(departments):
        start_index += 1
        ws['A' + str(start_index)] = index + 1
        ws['B' + str(start_index)] = department.title
        groups = department.group_set.all()
        students = Student.objects.filter(group__id__in=groups)
        classes = Class.objects.filter(group__id__in=groups, semester=semester)
        all_passes = Pass.objects.filter(student__id__in=students, date__year=year, date__gte=start_semester_date, date__lte=end_semester_date)
        fill_line(all_passes, students, classes, start_index, ws, additional, None, department)
        start_index = fill_department(year, start_semester_date, end_semester_date, groups, start_index, ws, additional, department)

    groups = Group.objects.filter(department__faculty=faculty)
    students = Student.objects.filter(group__id__in=groups)
    classes = Class.objects.filter(group__id__in=groups)
    all_passes = Pass.objects.filter(student__id__in=students, date__year=year, date__gte=start_semester_date, date__lt=end_semester_date)
    start_index += 2
    ws['B' + str(start_index)] = 'Всього по факультету:'
    fill_line(all_passes, students, classes, start_index, ws, additional)
    start_index = fill_department(year, start_semester_date, end_semester_date, groups, start_index, ws, additional)

    workbook.save(FILE_NAME_DESTINATION + 'faculty_reduction_per_semester.xlsx')
    return STATIC_URL + 'reductions/' + 'faculty_reduction_per_semester.xlsx'


def fill_line(all_passes, students, classes, start_index, ws, additional, curs=None, department=None, flag=None):
    hours = 0
    curs = curs or ''
    department = department or ''
    flag = flag or ''
    if not curs and department:
        hours = sum([int(i) for i in additional[str(department.id)].values()])
    elif not curs and not department:
        for department in Department.objects.all():
            hours += sum([int(i) for i in additional[str(department.id)].values()])
    elif curs and not department:
        for department in Department.objects.all():
            key = '-'.join([str(department.id), str(curs)])
            if flag:
                key += '-' + flag
            hours += int(additional[str(department.id)][key])
    else:
        key = '-'.join([str(department.id), str(curs)])
        if flag:
            key += '-' + flag
        hours = int(additional[str(department.id)][key])
    passes_by_type_pass = all_passes.filter(type='pass')
    passes_not_by_type_pass = all_passes.exclude(type='pass')
    ws['C' + str(start_index)] = hours
    ws['D' + str(start_index)] = students.count()
    ws['E' + str(start_index)] = all_passes.distinct('student').count()
    ws['F' + str(start_index)] = ((all_passes.distinct('student').count() / students.count()) * 100) if students.count() > 0 else 0
    ws['G' + str(start_index)] = passes_by_type_pass.distinct('student').count()
    ws['H' + str(start_index)] = ((passes_by_type_pass.distinct('student').count() / students.count()) * 100) if students.count() > 0 else 0
    ws['I' + str(start_index)] = passes_by_type_pass.count() * 2
    ws['J' + str(start_index)] = all_passes.count() * 2
    ws['K' + str(start_index)] = ((passes_by_type_pass.count() / (hours * students.count())) * 100) if classes.count() > 0 else 0
    ws['L' + str(start_index)] = ((all_passes.count() / (hours * students.count())) * 100) if classes.count() > 0 else 0
    return ws


def fill_department(year, start_date, end_date, groups, start_index, ws, additional, department=None):
    for curs in range(1, 6):
            start_index += 1
            ws['B' + str(start_index)] = str(curs) + ' курс'
            groups = groups.filter(yearStudy=curs)
            classes = Class.objects.filter(group__id__in=groups)
            students = Student.objects.filter(group__id__in=groups)
            all_passes = Pass.objects.filter(student__id__in=students, date__year=year, date__gte=start_date, date__lt=end_date)
            if curs == 5:
                fill_line(all_passes, students, classes, start_index, ws, additional, curs, department, 'spec')
            else:
                fill_line(all_passes, students, classes, start_index, ws, additional, curs, department)
            if curs == 3 or curs == 4:
                start_index += 1
                ws['B' + str(start_index)] = str(curs) + ' курс (скор. форма)'
                groups = groups.filter(yearStudy=curs, short_form=True)
                classes = Class.objects.filter(group__id__in=groups)
                students = Student.objects.filter(group__id__in=groups)
                all_passes = Pass.objects.filter(student__id__in=students, date__year=year, date__gte=start_date, date__lt=end_date)
                fill_line(all_passes, students, classes, start_index, ws, additional, curs, department, 'short')
            if curs == 5:
                start_index += 1
                ws['B' + str(start_index)] = str(curs) + ' курс (М)'
                groups = groups.filter(yearStudy=curs, master=True)
                classes = Class.objects.filter(group__id__in=groups)
                students = Student.objects.filter(group__id__in=groups)
                all_passes = Pass.objects.filter(student__id__in=students, date__year=year, date__gte=start_date, date__lt=end_date)
                fill_line(all_passes, students, classes, start_index, ws, additional, curs, department, 'master')

    return start_index